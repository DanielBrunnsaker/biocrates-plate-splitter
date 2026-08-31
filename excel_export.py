from __future__ import annotations
import io
from typing import List, Dict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

from algorithm import Plate, PLATE_ROWS, ALL_COLS, SAMPLE_POSITIONS

# ── Color palette ─────────────────────────────────────────────────────────────

STUDY_COLORS_HEX: List[str] = [
    "FF6B6B", "4ECDC4", "45B7D1", "96CEB4", "FFEAA7",
    "DDA0DD", "A8E6CF", "FFB7B2", "B8E0D2", "C7CEEA",
    "FFDAC1", "E2F0CB", "B5EAD7", "FF9AA2", "D4A5A5",
    "85C1E9", "A9DFBF", "F9E79F", "D7BDE2", "AED6F1",
]

CAL_COLOR       = "CCCCCC"
EMPTY_COLOR     = "FFFFFF"

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def get_study_color_map(study_ids: List[str]) -> Dict[str, str]:
    """Map each study_id to a hex color string (no '#')."""
    return {
        sid: STUDY_COLORS_HEX[i % len(STUDY_COLORS_HEX)]
        for i, sid in enumerate(study_ids)
    }


def export_plates_to_excel(plates: List[Plate], study_ids: List[str]) -> bytes:
    color_map = get_study_color_map(study_ids)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb, plates, color_map)
    for plate in plates:
        _write_plate_sheet(wb, plate, color_map)
    _write_sample_list(wb, plates)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_summary(wb: openpyxl.Workbook, plates: List[Plate], color_map: Dict[str, str]) -> None:
    ws = wb.create_sheet("Summary")

    ws["A1"] = "Plate Splitter — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    headers = ["Plate", "Samples", "Empty slots", "Studies"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for i, plate in enumerate(plates):
        ws_list = list(plate.wells.values())
        n_samples  = sum(1 for w in ws_list if w.well_type == "sample")
        n_empty    = sum(1 for w in ws_list if w.well_type == "empty")
        studies    = sorted({w.study_id for w in ws_list if w.well_type == "sample" and w.study_id})

        row = 4 + i
        ws.cell(row=row, column=1, value=f"Plate {plate.plate_number}").font = Font(bold=True)
        ws.cell(row=row, column=2, value=n_samples)
        ws.cell(row=row, column=3, value=n_empty)
        ws.cell(row=row, column=4, value=", ".join(studies))

    # Legend
    legend_row = 5 + len(plates)
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    col = 2
    for label, hex_color in [("Calibrators/QC", CAL_COLOR)]:
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        col += 1

    for sid, hex_color in color_map.items():
        c = ws.cell(row=legend_row, column=col, value=sid)
        c.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(sid) + 2)
        col += 1

    for j in range(1, 5):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.row_dimensions[3].height = 20


def _write_plate_sheet(wb: openpyxl.Workbook, plate: Plate, color_map: Dict[str, str]) -> None:
    ws = wb.create_sheet(f"Plate {plate.plate_number}")

    # Title row
    ws["A1"] = f"Plate {plate.plate_number}"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 22

    # Column headers (row 2: blank then 1-12)
    ws.cell(row=2, column=1, value="")
    ws.column_dimensions["A"].width = 4
    for col_num in ALL_COLS:
        c = ws.cell(row=2, column=col_num + 1, value=col_num)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 13

    # Row headers + well contents
    for r_idx, row_ltr in enumerate(PLATE_ROWS):
        excel_row = r_idx + 3
        # Row label
        c = ws.cell(row=excel_row, column=1, value=row_ltr)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[excel_row].height = 32

        for col_num in ALL_COLS:
            w = plate.wells[(row_ltr, col_num)]
            cell = ws.cell(row=excel_row, column=col_num + 1)

            if w.well_type == "calibrator":
                text = w.label or "CAL"
                color = CAL_COLOR
            elif w.well_type == "empty":
                text = ""
                color = EMPTY_COLOR
            else:
                text = w.sample_id or ""
                color = color_map.get(w.study_id or "", "EEEEEE")

            cell.value = text
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(size=8)

    # Legend below plate
    legend_row = 3 + len(PLATE_ROWS) + 1
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    col = 2
    for label, hex_c in [("Calibrators/QC", CAL_COLOR), ("Empty", EMPTY_COLOR)]:
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        col += 1

    studies_on_plate = sorted({
        w.study_id for w in plate.wells.values()
        if w.well_type == "sample" and w.study_id
    })
    for sid in studies_on_plate:
        hex_c = color_map.get(sid, "EEEEEE")
        c = ws.cell(row=legend_row, column=col, value=sid)
        c.fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        col += 1


def _write_sample_list(wb: openpyxl.Workbook, plates: List[Plate]) -> None:
    """Full sample → plate mapping sheet."""
    ws = wb.create_sheet("Sample List")
    headers = [
        "Plate", "Well", "Row", "Column", "Sample ID", "Study ID", "Well Type",
        "Repeat Pair", "Study Factor 1", "Study Factor 2", "Biological Matrix",
        "Sample Description", "Sample Alias",
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 14

    row_idx = 2
    for plate in plates:
        df = plate.to_dataframe()
        df = df[df["well_type"] == "sample"].sort_values(["col", "row"])
        for _, record in df.iterrows():
            ws.cell(row=row_idx, column=1, value=record["plate"])
            ws.cell(row=row_idx, column=2, value=record["well"])
            ws.cell(row=row_idx, column=3, value=record["row"])
            ws.cell(row=row_idx, column=4, value=record["col"])
            ws.cell(row=row_idx, column=5, value=record["sample_id"])
            ws.cell(row=row_idx, column=6, value=record["study_id"])
            ws.cell(row=row_idx, column=7, value=record["well_type"])
            ws.cell(row=row_idx, column=8, value=record["repeat_pair"])
            ws.cell(row=row_idx, column=9, value=record.get("factor_1", ""))
            ws.cell(row=row_idx, column=10, value=record.get("factor_2", ""))
            ws.cell(row=row_idx, column=11, value=record.get("material", ""))
            ws.cell(row=row_idx, column=12, value=record.get("sample_description", ""))
            ws.cell(row=row_idx, column=13, value=record.get("sample_alias", ""))
            row_idx += 1
